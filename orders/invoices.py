"""Generate a downloadable «Счёт на оплату» (XLSX) for an order."""

from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from about.models import CompanyDetails

_MONTHS = [
    "", "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]


def _money(value):
    return f"{Decimal(value):,.2f}".replace(",", " ").replace(".", ",")


def _rate_label(rate):
    return format(Decimal(rate).normalize(), "f").replace(".", ",")


def _amount_in_words(total):
    try:
        from num2words import num2words

        words = num2words(float(total), lang="ru", to="currency", currency="RUB")
        return words[:1].upper() + words[1:]
    except Exception:
        return ""


def _short_name(full):
    parts = (full or "").split()
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1][0]}. {parts[2][0]}."
    if len(parts) == 2:
        return f"{parts[0]} {parts[1][0]}."
    return full or ""


def _date_str(dt):
    d = timezone.localtime(dt)
    return f"{d.day} {_MONTHS[d.month]} {d.year} г."


def build_invoice_xlsx(order):
    co = CompanyDetails.load()
    user = order.user
    buyer = user.company_name or user.get_full_name() or user.username

    wb = Workbook()
    ws = wb.active
    ws.title = "Счёт"
    for col, w in {"A": 5, "B": 42, "C": 16, "D": 9, "E": 7, "F": 14, "G": 16}.items():
        ws.column_dimensions[col].width = w

    # A4 portrait, scaled to fit the page width — convenient to print.
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)

    base = Font(name="Arial", size=9)
    bold = Font(name="Arial", size=9, bold=True)
    title = Font(name="Arial", size=12, bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    R = Alignment(horizontal="right", vertical="center")

    r = 1

    def wide(text, font=base, align=L, height=None):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(r, 1, text)
        cell.font, cell.alignment = font, align
        if height:
            ws.row_dimensions[r].height = height
        r += 1

    # Seller / bank block
    wide(f"ИНН {co.inn}    КПП {co.kpp}", base)
    wide(co.name, bold)
    wide(f"Банк получателя: {co.bank}", base, height=30)
    wide(f"БИК: {co.bank_bic}    Корр. счёт: {co.corr_account}", base)
    wide(f"Расчётный счёт: {co.settlement_account}", base)
    r += 1

    # Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    t = ws.cell(r, 1, f"Счёт на оплату № {order.invoice_number} от {_date_str(order.created_at)}")
    t.font, t.alignment = title, Alignment(horizontal="center")
    r += 2

    # Parties
    wide(
        f"Поставщик (Исполнитель): {co.name}, ИНН {co.inn}, КПП {co.kpp}, {co.address}",
        base, height=45,
    )
    buyer_line = (
        f"Покупатель (Заказчик): {buyer}, ИНН {user.inn}, КПП {user.kpp}, {user.address}"
    )
    if user.phone:
        buyer_line += f", тел.: {user.phone}"
    wide(buyer_line, base, height=45)
    r += 1

    # Items table
    headers = ["№", "Товар (Услуга)", "Код", "Кол-во", "Ед.", "Цена", "Сумма"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(r, i, h)
        cell.font, cell.alignment, cell.border = bold, C, border
        cell.fill = PatternFill("solid", fgColor="EFEFEF")
    ws.row_dimensions[r].height = 28
    r += 1

    vat_by_rate = {}
    items = list(order.items.select_related("product").all())
    for idx, item in enumerate(items, start=1):
        rate = item.product.vat_rate if item.product else Decimal(22)
        sub = item.subtotal or Decimal(0)
        if rate:
            vat_by_rate[rate] = vat_by_rate.get(rate, Decimal(0)) + sub * rate / (100 + rate)
        code = item.product.article if item.product else item.sku
        values = [idx, item.name, code, item.quantity, "шт", float(item.price), float(sub)]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(r, i, v)
            cell.font, cell.border = base, border
            if i == 2:
                cell.alignment = L
            elif i in (6, 7):
                cell.alignment, cell.number_format = R, "#,##0.00"
            else:
                cell.alignment = C
        ws.row_dimensions[r].height = 15 * max(1, len(item.name) // 50 + 1)
        r += 1

    # Totals
    def total_line(label, value, strong=False):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        lc = ws.cell(r, 1, label)
        lc.font, lc.alignment = (bold if strong else base), R
        vc = ws.cell(r, 7, value)
        vc.font, vc.alignment, vc.number_format = (bold if strong else base), R, "#,##0.00"
        r += 1

    total_line("Итого:", float(order.total), True)
    for rate in sorted(vat_by_rate):
        total_line(f"В том числе НДС {_rate_label(rate)}%:", float(round(vat_by_rate[rate], 2)))
    total_line("Всего к оплате:", float(order.total), True)
    r += 1

    wide(f"Всего наименований {len(items)}, на сумму {_money(order.total)} руб.", bold)
    words = _amount_in_words(order.total)
    if words:
        wide(words, bold, height=24)
    r += 1

    for note in (
        "Оплата данного счёта означает согласие с условиями поставки товара.",
        "Товар отпускается по факту прихода денег на р/с Поставщика.",
    ):
        wide(note, base)
    r += 1

    wide(f"Руководитель ______________________   {_short_name(co.director)}", base)
    wide(f"Бухгалтер ______________________   {_short_name(co.director)}", base)
    wide("М.П.", base)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
