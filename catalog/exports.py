"""Export the whole catalog to an .xlsx workbook (admin-only).

The layout mirrors :mod:`catalog.imports`, so a downloaded file can be edited
and uploaded straight back: every product column below is one the import
recognises, price columns are named after their PriceType, and the remaining
columns after each Warehouse (holding stock quantities). The trailing columns
in EXTRA_COLUMNS are informational — the import skips them by name rather than
mistaking them for warehouses.
"""

from io import BytesIO

import openpyxl
from django.utils import timezone
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from warehouses.models import Warehouse

from .models import PriceType, Product

EXCEL_MAX_CHARS = 32767


def _clean(value):
    """Excel rejects control characters and caps a cell at 32767 chars."""
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)[:EXCEL_MAX_CHARS]


def _moment(value):
    return timezone.localtime(value).strftime("%d.%m.%Y %H:%M") if value else ""


# Columns the import understands — see catalog.imports.FIELD_ALIASES.
PRODUCT_COLUMNS = [
    ("sku", lambda p: p.sku),
    ("brand", lambda p: p.brand.name if p.brand_id else ""),
    ("name", lambda p: p.name),
    ("article", lambda p: p.article),
    ("manufacturer_number", lambda p: p.manufacturer_number),
    ("category", lambda p: p.category.name if p.category_id else ""),
    ("subcategory", lambda p: p.subcategory.name if p.subcategory_id else ""),
    ("model_product", lambda p: p.model_product.name if p.model_product_id else ""),
    ("weight", lambda p: p.weight),
    ("weight_unit", lambda p: p.weight_unit),
    ("volume", lambda p: p.volume),
    ("volume_unit", lambda p: p.volume_unit),
    ("pack_quantity", lambda p: p.pack_quantity),
    ("viscosity", lambda p: p.viscosity),
    ("description", lambda p: p.description),
    ("price", lambda p: p.price),
    ("vat_rate", lambda p: p.vat_rate),
]

# Informational only; catalog.imports ignores these headers on the way back in.
EXTRA_COLUMNS = [
    ("is_active", lambda p: p.is_active),
    ("slug", lambda p: p.slug),
    ("picture", lambda p: p.picture.name if p.picture else ""),
    ("created_at", lambda p: _moment(p.created_at)),
    ("updated_at", lambda p: _moment(p.updated_at)),
]


def build_catalog_xlsx():
    """Return the full catalog (every product, every attribute) as .xlsx bytes."""
    price_types = list(PriceType.objects.order_by("name"))
    warehouses = list(Warehouse.objects.order_by("name"))
    products = (
        Product.objects.select_related(
            "brand", "category", "subcategory", "model_product"
        )
        .prefetch_related("prices", "stocks")
        .order_by("brand__name", "name")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог"

    header = (
        [name for name, _ in PRODUCT_COLUMNS]
        + [pt.name for pt in price_types]
        + [w.name for w in warehouses]
        + [name for name, _ in EXTRA_COLUMNS]
    )
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for product in products:
        prices = {p.price_type_id: p.price for p in product.prices.all()}
        stock = {s.warehouse_id: s.quantity for s in product.stocks.all()}
        row = [_clean(get(product)) for _, get in PRODUCT_COLUMNS]
        row += [prices.get(pt.pk) for pt in price_types]
        row += [stock.get(w.pk, 0) for w in warehouses]
        row += [_clean(get(product)) for _, get in EXTRA_COLUMNS]
        ws.append(row)

    for idx, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = (
            60 if name == "description" else max(12, min(len(name) + 4, 24))
        )

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
