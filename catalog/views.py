import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from urllib.parse import quote

import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import DecimalField, F, IntegerField, OuterRef, Q, Subquery, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from warehouses.availability import (
    annotate_availability,
    own_warehouse_ids,
    warehouse_breakdown,
)
from warehouses.selection import get_current_warehouse

from .models import BannerSlide, Brand, Category, Product, ProductPrice, SubCategory
from .pricing import price_type_for_user
from .visibility import visible_brands_q, visible_products_q


def _viscosity_key(value):
    """Sort viscosities numerically (0W20, 5W30, 10W40) rather than as text."""
    match = re.match(r"(\d+)W(\d+)", value)
    return (int(match.group(1)), int(match.group(2))) if match else (10**6, 10**6)


def _with_effective_price(products, price_type):
    """Annotate each product with `effective_price` for the given price type,
    falling back to the product's base `price` when no per-type price exists."""
    if price_type is not None:
        price_sub = ProductPrice.objects.filter(
            product=OuterRef("pk"), price_type=price_type
        ).values("price")[:1]
        return products.annotate(
            effective_price=Coalesce(
                Subquery(
                    price_sub,
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                F("price"),
            )
        )
    return products.annotate(effective_price=F("price"))


RECENT_KEY = "recently_viewed"
RECENT_STORE = 20  # product ids kept in the session
RECENT_SHOW = 6  # products shown in the block


def _remember_viewed(request, pk):
    """Record a product as recently viewed (newest first, deduped, capped)."""
    ids = [i for i in request.session.get(RECENT_KEY, []) if i != pk]
    ids.insert(0, pk)
    request.session[RECENT_KEY] = ids[:RECENT_STORE]


def _recently_viewed(
    request, price_type, warehouse_ids, exclude_pk=None, limit=RECENT_SHOW
):
    """Recently viewed products (session-based), newest first, price-annotated."""
    ids = [i for i in request.session.get(RECENT_KEY, []) if i != exclude_pk]
    if not ids:
        return []
    products = _with_effective_price(
        Product.objects.filter(pk__in=ids, is_active=True)
        .filter(visible_products_q(warehouse_ids))
        .select_related("brand"),
        price_type,
    )
    by_id = {p.pk: p for p in products}
    result = []
    for i in ids:
        product = by_id.get(i)
        if product is not None:
            result.append(product)
        if len(result) >= limit:
            break
    return result


PER_PAGE_OPTIONS = (12, 50, 100)


def _favorite_ids(user):
    """Product ids in the user's wishlist (empty set for guests)."""
    if not getattr(user, "is_authenticated", False):
        return set()
    return set(user.favorites.values_list("product_id", flat=True))


def product_list(request):
    # Public page: anyone may browse. Stock (across all warehouses) is revealed
    # to any logged-in buyer; `warehouse` is their own warehouse for ordering.
    warehouse = get_current_warehouse(request)
    show_stock = request.user.is_authenticated
    own_ids = own_warehouse_ids(request.user)
    products = (
        Product.objects.filter(is_active=True)
        .filter(visible_products_q(own_ids))
        .select_related("brand", "category", "subcategory")
    )

    query = request.GET.get("q", "").strip()
    brand_slug = request.GET.get("brand", "")
    category_id = request.GET.get("category", "")
    subcategory_id = request.GET.get("subcategory", "")
    volume_value = request.GET.get("volume", "")
    viscosity_value = request.GET.get("viscosity", "")
    in_stock = request.GET.get("in_stock") == "1"

    if query:
        products = products.filter(
            Q(name__icontains=query)
            | Q(sku__icontains=query)
            | Q(article__icontains=query)
            | Q(manufacturer_number__icontains=query)
            | Q(brand__name__icontains=query)
            | Q(description__icontains=query)
        )
    # Brand → categories cascade. Which categories a brand offers is set in the
    # admin (Brand.categories); empty means "all categories".
    brand_obj = Brand.objects.filter(slug=brand_slug).first() if brand_slug else None
    categories_qs = Category.objects.prefetch_related("subcategories")
    if brand_obj and brand_obj.categories.exists():
        categories_qs = categories_qs.filter(brands=brand_obj)
    categories = list(categories_qs)
    allowed_cat_ids = {c.id for c in categories}

    # Drop selections that no longer fit the cascade (e.g. a category not sold by
    # the chosen brand, or a subcategory outside the chosen category).
    if category_id.isdigit() and int(category_id) not in allowed_cat_ids:
        category_id = ""
    if subcategory_id.isdigit():
        sub = SubCategory.objects.filter(pk=subcategory_id).first()
        if sub is None:
            subcategory_id = ""
        elif category_id.isdigit():
            if str(sub.category_id) != category_id:
                subcategory_id = ""
        elif sub.category_id not in allowed_cat_ids:
            subcategory_id = ""

    if brand_slug:
        products = products.filter(brand__slug=brand_slug)
    if category_id.isdigit():
        products = products.filter(category_id=category_id)
    if subcategory_id.isdigit():
        products = products.filter(subcategory_id=subcategory_id)

    # Volumes and viscosities available for the current brand/category/subcategory.
    # Distinct (volume, unit) pairs so the filter shows the real unit per size
    # (e.g. "450 мл" and "5 л"), now that units are per-product.
    volumes = [
        {"token": f"{v}|{u}", "value": v, "unit": u}
        for v, u in products.filter(volume__isnull=False)
        .values_list("volume", "volume_unit")
        .distinct()
        .order_by("volume_unit", "volume")
    ]
    viscosities = sorted(
        set(products.exclude(viscosity="").values_list("viscosity", flat=True)),
        key=_viscosity_key,
    )
    # Apply the volume filter only if that (volume, unit) option is available.
    if volume_value:
        if volume_value in {opt["token"] for opt in volumes}:
            raw_value, _, raw_unit = volume_value.partition("|")
            products = products.filter(volume=Decimal(raw_value), volume_unit=raw_unit)
        else:
            volume_value = ""
    # Apply the viscosity filter only if that viscosity is actually available.
    if viscosity_value:
        if viscosity_value in viscosities:
            products = products.filter(viscosity=viscosity_value)
        else:
            viscosity_value = ""

    # Buyers see stock across all warehouses: own warehouses (immediate) plus
    # every other warehouse (7-day delivery). Availability = the total.
    if show_stock:
        products = annotate_availability(products, own_ids)

    # "Show in stock only" — available anywhere (own or 7-day).
    if in_stock and show_stock:
        products = products.filter(total_qty__gt=0)

    # Prices shown depend on the viewer's price type (guests → «Розничные»).
    price_type = price_type_for_user(request.user)
    products = _with_effective_price(products, price_type)

    # Explicit ordering: the Sum annotations above drop the model's default
    # ordering, which would make pagination inconsistent.
    products = products.order_by("name")
    try:
        per_page = int(request.GET.get("per_page", PER_PAGE_OPTIONS[0]))
    except (TypeError, ValueError):
        per_page = PER_PAGE_OPTIONS[0]
    if per_page not in PER_PAGE_OPTIONS:
        per_page = PER_PAGE_OPTIONS[0]
    paginator = Paginator(products, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Show the rotating banner only on the clean landing page (no search/filter).
    is_landing = not any(
        [query, brand_slug, category_id, subcategory_id, volume_value, viscosity_value, in_stock, request.GET.get("page")]
    )
    banner_slides = (
        BannerSlide.objects.filter(is_active=True)[:7] if is_landing else []
    )

    context = {
        "banner_slides": banner_slides,
        "page_obj": page_obj,
        "warehouse": warehouse,
        "show_stock": show_stock,
        "brands": Brand.objects.filter(visible_brands_q(own_ids)),
        "categories": categories,
        "volumes": volumes,
        "viscosities": viscosities,
        "query": query,
        "selected_brand": brand_slug,
        "selected_category": category_id,
        "selected_subcategory": subcategory_id,
        "selected_volume": volume_value,
        "selected_viscosity": viscosity_value,
        "selected_in_stock": "1" if in_stock else "",
        "price_type": price_type,
        "per_page": per_page,
        "per_page_options": PER_PAGE_OPTIONS,
        "favorite_ids": _favorite_ids(request.user),
        "recently_viewed": _recently_viewed(request, price_type, own_ids),
    }
    return render(request, "catalog/product_list.html", context)


def product_detail(request, pk):
    warehouse = get_current_warehouse(request)
    own_ids = own_warehouse_ids(request.user)
    # A brand hidden at every warehouse the user holds is out of the catalog, so
    # its products must be unreachable by direct URL too.
    product = get_object_or_404(
        Product.objects.filter(visible_products_q(own_ids)).select_related(
            "brand", "category", "subcategory", "model_product"
        ),
        pk=pk,
        is_active=True,
    )
    show_stock = request.user.is_authenticated
    # Stock per warehouse: own warehouses (immediate) + others (7-day delivery).
    warehouse_stock = warehouse_breakdown(product, own_ids) if show_stock else []
    total_available = sum(row["qty"] for row in warehouse_stock)
    # Price shown depends on the viewer's price type (guests → «Розничные»).
    price_type = price_type_for_user(request.user)
    price = product.price_for(price_type)
    # Other products of the same model (different volume or weight), ordered by
    # volume then weight so weight-only variants sort by weight.
    variants = []
    if product.model_product_id:
        variants = list(
            _with_effective_price(
                Product.objects.filter(
                    model_product_id=product.model_product_id, is_active=True
                ),
                price_type,
            ).order_by("volume", "weight", "name")
        )

    # Recently viewed (excluding this product), then record this view.
    recently_viewed = _recently_viewed(
        request, price_type, own_ids, exclude_pk=product.pk
    )
    _remember_viewed(request, product.pk)

    context = {
        "product": product,
        "warehouse": warehouse,
        "show_stock": show_stock,
        "warehouse_stock": warehouse_stock,
        "total_available": total_available,
        "variants": variants,
        "price": price,
        "price_type": price_type,
        "favorite_ids": _favorite_ids(request.user),
        "recently_viewed": recently_viewed,
    }
    return render(request, "catalog/product_detail.html", context)


def _filtered_products(request):
    """Active products with the catalog GET-param filters applied."""
    products = (
        Product.objects.filter(is_active=True)
        .filter(visible_products_q(own_warehouse_ids(request.user)))
        .select_related("brand", "category", "subcategory")
    )
    query = request.GET.get("q", "").strip()
    if query:
        products = products.filter(
            Q(name__icontains=query)
            | Q(sku__icontains=query)
            | Q(article__icontains=query)
            | Q(manufacturer_number__icontains=query)
            | Q(brand__name__icontains=query)
            | Q(description__icontains=query)
        )
    brand_slug = request.GET.get("brand", "")
    if brand_slug:
        products = products.filter(brand__slug=brand_slug)
    category_id = request.GET.get("category", "")
    if category_id.isdigit():
        products = products.filter(category_id=category_id)
    subcategory_id = request.GET.get("subcategory", "")
    if subcategory_id.isdigit():
        products = products.filter(subcategory_id=subcategory_id)
    volume_value = request.GET.get("volume", "")
    if volume_value and "|" in volume_value:
        raw_value, _, raw_unit = volume_value.partition("|")
        try:
            products = products.filter(volume=Decimal(raw_value), volume_unit=raw_unit)
        except (InvalidOperation, ValueError):
            pass
    viscosity_value = request.GET.get("viscosity", "")
    if viscosity_value:
        products = products.filter(viscosity=viscosity_value)
    return products


@login_required
def product_price_xlsx(request):
    """Download the (optionally filtered) catalog as an Excel price list.

    Columns: Наименование, Артикул, Бренд, Категория, Подкатегория,
    Количество (stock in the buyer's own warehouse), Цена (their price type).
    """
    own_ids = own_warehouse_ids(request.user)
    price_type = price_type_for_user(request.user)
    products = _filtered_products(request)
    products = annotate_availability(products, own_ids)
    products = _with_effective_price(products, price_type)
    if request.GET.get("in_stock") == "1":
        products = products.filter(total_qty__gt=0)
    products = products.order_by(
        "brand__name", "category__name", "subcategory__name", "name"
    )

    today = timezone.localdate().strftime("%d.%m.%Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс"
    headers = [
        "Наименование", "Артикул", "Бренд", "Категория",
        "Подкатегория", "Количество", "Цена",
    ]
    # Title row spanning all columns.
    ws.append([f"Прайс Автомеханика-Сибирь по состоянию на {today}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    # Column headers on the second row.
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A3"

    for p in products.iterator():
        ws.append([
            p.name,
            p.article,
            p.brand.name if p.brand_id else "",
            p.category.name if p.category_id else "",
            p.subcategory.name if p.subcategory_id else "",
            p.own_qty,
            float(p.effective_price or 0),
        ])
    for i, w in enumerate([45, 16, 18, 20, 20, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    resp = HttpResponse(
        bio.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    fname = f"прайс_амс_{today}.xlsx"
    resp["Content-Disposition"] = (
        f"attachment; filename=\"price.xlsx\"; filename*=UTF-8''{quote(fname)}"
    )
    return resp
